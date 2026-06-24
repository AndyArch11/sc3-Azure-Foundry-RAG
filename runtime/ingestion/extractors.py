from __future__ import annotations

import logging
import warnings
from pathlib import Path

from .models import SourceDocument

SUPPORTED_EXTENSIONS = {".pdf", ".xlsx", ".xlsm", ".xltx", ".xltm"}
logger = logging.getLogger(__name__)


def _extract_pdf_text_ocr(path: Path) -> str:
    """Run extract pdf text ocr."""
    try:
        import pypdfium2 as pdfium
        import pytesseract
    except ImportError as exc:
        raise RuntimeError("pypdfium2 and pytesseract are required for local OCR fallback") from exc

    pages: list[str] = []
    document = pdfium.PdfDocument(str(path))
    try:
        for idx in range(len(document)):
            page = document[idx]
            # Scale up for better OCR quality on scanned PDFs.
            image = page.render(scale=2).to_pil()
            text = pytesseract.image_to_string(image) or ""
            if text.strip():
                pages.append(text)
    finally:
        document.close()

    return "\n".join(pages)


def _extract_pdf_text(path: Path, *, enable_ocr: bool = False, min_text_chars: int = 80) -> str:
    """Run extract pdf text."""
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise RuntimeError("pypdf is required for PDF ingestion") from exc

    reader = PdfReader(str(path))
    pages: list[str] = []
    for page in reader.pages:
        pages.append(page.extract_text() or "")

    extracted = "\n".join(pages)
    if not enable_ocr:
        return extracted

    # For scanned/image-heavy PDFs, pypdf text can be sparse. Trigger OCR only
    # when extracted text is below a practical threshold.
    if len(extracted.strip()) >= max(0, min_text_chars):
        return extracted

    try:
        ocr_text = _extract_pdf_text_ocr(path)
    except RuntimeError as exc:
        logger.warning("Local OCR unavailable for %s: %s", path, exc)
        return extracted

    if not ocr_text.strip():
        return extracted
    if not extracted.strip():
        return ocr_text
    return f"{extracted}\n{ocr_text}"


def _extract_excel_text(path: Path) -> str:
    """Run extract excel text."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for Excel ingestion") from exc

    with warnings.catch_warnings():
        # Some vendor workbooks include optional OOXML extensions that openpyxl
        # ignores safely; suppress the noisy warning in ingestion logs.
        warnings.filterwarnings(
            "ignore",
            message="Unknown extension is not supported and will be removed",
            category=UserWarning,
            module=r"openpyxl\\..*",
        )
        workbook = load_workbook(path, data_only=True, read_only=True)
    lines: list[str] = []
    for sheet in workbook.worksheets:
        lines.append(f"# Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
            if values:
                lines.append(" | ".join(values))
    return "\n".join(lines)


def extract_source_document(
    path: Path,
    *,
    enable_ocr: bool = False,
    ocr_min_text_chars: int = 80,
) -> SourceDocument:
    """Run extract source document."""
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        text = _extract_pdf_text(
            path,
            enable_ocr=enable_ocr,
            min_text_chars=ocr_min_text_chars,
        )
        source_type = "pdf"
    elif suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        text = _extract_excel_text(path)
        source_type = "excel"
    else:
        raise ValueError(f"Unsupported file type for ingestion: {path.name}")

    return SourceDocument(source_path=str(path), source_type=source_type, text=text)


def discover_supported_files(input_dir: Path) -> list[Path]:
    """Run discover supported files."""
    files = [
        p for p in input_dir.rglob("*") if p.is_file() and p.suffix.lower() in SUPPORTED_EXTENSIONS
    ]
    return sorted(files)
