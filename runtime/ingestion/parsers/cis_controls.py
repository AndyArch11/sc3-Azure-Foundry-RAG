"""Parser for CIS Critical Security Controls v8 using local sample workbook and PDF.

The workbook provides the structured safeguard rows. The PDF provides narrative
guidance per control (overview, why critical, and procedures/tools).
"""

from __future__ import annotations

import io
import logging
import re
from pathlib import Path
from typing import Any, Dict, List

from .base import BaseParser, RequirementRecord, filter_keywords, keywordise_values

logger = logging.getLogger(__name__)

try:
    from pypdf import PdfReader as _PdfReader
except ImportError:
    _PdfReader = None  # type: ignore[misc,assignment]

FRAMEWORK = "CIS Controls"
FRAMEWORK_VERSION = "v8"
EFFECTIVE_DATE = "May 2021"
JURISDICTION = "Global"
SOURCE_URI = "https://www.cisecurity.org/controls/v8"

_DEFAULT_WORKBOOK_PATH = (
    Path(__file__).resolve().parents[2] / "samples" / "CIS_Controls_Version_8.xlsx"
)
_DEFAULT_PDF_PATH = (
    Path(__file__).resolve().parents[2]
    / "samples"
    / "CIS_Controls__v8__Critical_Security_Controls__2023_08.pdf"
)


def _slugify(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", text.lower()).strip("_")


def _maturity_level_from_igs(ig1: Any, ig2: Any, ig3: Any) -> int | None:
    """Return the minimum implementation group (1-3) marked with 'x'."""
    for level, value in ((1, ig1), (2, ig2), (3, ig3)):
        if str(value or "").strip().lower() == "x":
            return level
    return None


def _normalise_pdf_text(text: str) -> str:
    text = text.replace("\xa0", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def _normalise_guidance_text(text: str) -> str:
    """Flatten guidance text to avoid hard line-break artifacts from PDF extraction."""
    # Repair discretionary hyphenation introduced by line-wrapped PDF extraction.
    text = text.replace("-\n", "-")
    # Collapse all remaining whitespace/newlines into single spaces for stable downstream rendering.
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _extract_section(text: str, start_label: str, end_labels: list[str]) -> str:
    pattern = re.escape(start_label) + r"\s*(.*?)"
    if end_labels:
        pattern += r"(?=" + "|".join(re.escape(label) for label in end_labels) + r"|\Z)"
    match = re.search(pattern, text, re.DOTALL)
    if not match:
        return ""
    return _normalise_pdf_text(match.group(1))


def _build_control_guidance_map(pdf_path: Path) -> dict[str, str]:
    if _PdfReader is None:
        raise RuntimeError(
            "pypdf is required for CIS Controls PDF parsing. Install with: pip install pypdf"
        )

    reader = _PdfReader(str(pdf_path))
    sections_by_control: dict[str, list[str]] = {}
    current_control = ""

    for page in reader.pages:
        text = _normalise_pdf_text(page.extract_text() or "")
        if not text:
            continue
        start_match = re.search(r"(?m)^\s*(\d{2})\s*$", text)
        if start_match and "OVERVIEW" in text:
            current_control = str(int(start_match.group(1)))
            sections_by_control.setdefault(current_control, [])
        if current_control:
            sections_by_control.setdefault(current_control, []).append(text)

    guidance_map: dict[str, str] = {}
    for control_number, parts in sections_by_control.items():
        combined = "\n\n".join(parts)
        overview = _extract_section(combined, "OVERVIEW", ["Why is this Control critical?"])
        why_critical = _extract_section(
            combined,
            "Why is this Control critical?",
            ["Procedures and tools", "Safeguards", "CONTROL"],
        )
        procedures = _extract_section(combined, "Procedures and tools", ["Safeguards", "CONTROL"])

        guidance_parts = []
        if overview:
            guidance_parts.append(f"Overview: {_normalise_guidance_text(overview)}")
        if why_critical:
            guidance_parts.append(f"Why critical: {_normalise_guidance_text(why_critical)}")
        if procedures:
            guidance_parts.append(f"Procedures and tools: {_normalise_guidance_text(procedures)}")
        if guidance_parts:
            guidance_map[control_number] = " ".join(guidance_parts)

    return guidance_map


class CisControlsParser(BaseParser):
    def __init__(
        self,
        workbook_path: str | Path = _DEFAULT_WORKBOOK_PATH,
        pdf_path: str | Path = _DEFAULT_PDF_PATH,
        **_kwargs: Any,
    ) -> None:
        self._workbook_path = Path(workbook_path)
        self._pdf_path = Path(pdf_path)

    def parse(self) -> List[RequirementRecord]:
        try:
            import openpyxl  # noqa: PLC0415
        except ImportError as exc:
            raise RuntimeError(
                "openpyxl is required for CIS Controls workbook parsing. Install with: pip install openpyxl"
            ) from exc

        if not self._workbook_path.exists():
            raise RuntimeError(f"CIS workbook not found: {self._workbook_path}")
        if not self._pdf_path.exists():
            raise RuntimeError(f"CIS PDF not found: {self._pdf_path}")

        guidance_map = _build_control_guidance_map(self._pdf_path)

        workbook = openpyxl.load_workbook(self._workbook_path, read_only=True, data_only=True)
        if "Controls V8" not in workbook.sheetnames:
            raise RuntimeError("CIS workbook missing expected 'Controls V8' sheet")
        sheet = workbook["Controls V8"]

        records: list[RequirementRecord] = []
        control_titles: dict[str, str] = {}
        control_descriptions: dict[str, str] = {}

        for row in sheet.iter_rows(min_row=2, values_only=True):
            (
                control_raw,
                safeguard_raw,
                asset_type_raw,
                security_function_raw,
                title_raw,
                description_raw,
                ig1_raw,
                ig2_raw,
                ig3_raw,
                *_rest,
            ) = row
            control_number = str(control_raw or "").strip()
            safeguard = str(safeguard_raw or "").strip()
            title = str(title_raw or "").strip()
            description = str(description_raw or "").strip()

            if control_number and not safeguard:
                control_key = str(int(control_number))
                control_titles[control_key] = title
                control_descriptions[control_key] = description
                continue

            if not control_number or not safeguard or not title or not description:
                continue

            control_key = str(int(control_number))
            control_title = control_titles.get(control_key, "")
            asset_type = str(asset_type_raw or "").strip()
            security_function = str(security_function_raw or "").strip(" -")
            implementation_groups = [
                group_name
                for group_name, value in (("ig1", ig1_raw), ("ig2", ig2_raw), ("ig3", ig3_raw))
                if str(value or "").strip().lower() == "x"
            ]
            maturity_level = _maturity_level_from_igs(ig1_raw, ig2_raw, ig3_raw)

            guidance_text = guidance_map.get(control_key) or control_descriptions.get(
                control_key, ""
            )
            keyword_values = [
                FRAMEWORK,
                FRAMEWORK_VERSION,
                control_title,
                title,
                asset_type,
                security_function,
                safeguard,
                *implementation_groups,
            ]
            keywords = keywordise_values(*keyword_values)
            keywords.extend(implementation_groups)
            keywords = filter_keywords(keywords)

            records.append(
                RequirementRecord(
                    requirement_id=f"CISv8-{safeguard.replace('.', '_')}",
                    framework=FRAMEWORK,
                    framework_version=FRAMEWORK_VERSION,
                    control_family=control_title or f"Control {control_key}",
                    maturity_level=maturity_level,
                    requirement_text=description,
                    guidance_text=guidance_text,
                    keywords=keywords,
                    source_uri=SOURCE_URI,
                    source_section=f"Control {int(control_key):02d} > Safeguard {safeguard}",
                    effective_date=EFFECTIVE_DATE,
                    jurisdiction_or_scope=JURISDICTION,
                )
            )

        logger.info("CIS Controls: parsed %d safeguards", len(records))
        return records
