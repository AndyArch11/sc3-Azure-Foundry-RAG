"""Parser for the Australian Energy Sector Cyber Security Framework (AESCSF) v2.

Source: AESCSF v2 Core workbook published by AEMO.
The workbook is fetched at parse time from the AEMO website.

The core workbook is a flat table with one row per practice, including domain,
objective, practice text, guidance text, MIL, and security profile.

One ``RequirementRecord`` is emitted per practice row.

Usage::

    from runtime.ingestion.parsers.aescsf import AescsfParser
    records = AescsfParser().parse()
    print(len(records))  # 354
"""

from __future__ import annotations

import io
import logging
import re
from typing import Dict, List, Optional

import requests  # type: ignore[import-untyped]

from runtime.outbound_instrumentation import request_with_instrumentation

from .base import BaseParser, RequirementRecord, filter_keywords

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

FRAMEWORK = "AESCSF"
FRAMEWORK_VERSION = "v2"
EFFECTIVE_DATE = "2023"
JURISDICTION = "Australia"

SOURCE_URI = (
    "https://www.aemo.com.au/initiatives/major-programs/cyber-security/"
    "aescsf-framework-and-resources"
)

_CORE_WORKBOOK_URL = (
    "https://www.aemo.com.au/-/media/files/initiatives/cyber-security/aescsf/2023/"
    "the-aescsf-v2-core.xlsx"
    "?rev=4375ddea4d394bee8b5c9bb7eb7fcbde&sc_lang=en"
)

_TOOLKIT_URL = (
    "https://www.aemo.com.au/-/media/files/initiatives/cyber-security/aescsf/2023/"
    "v2-aescsf-toolkit-version-v1-1.xlsx"
    "?rev=c918e8c7c8674376b877eded259eb4d7&sc_lang=en"
)

# Human-readable domain names keyed by the uppercase domain slug
_DOMAIN_NAMES: Dict[str, str] = {
    "ACCESS": "Access Management",
    "ARCHITECTURE": "Cybersecurity Architecture",
    "ASSET": "Asset, Change and Configuration Management",
    "PRIVACY": "Privacy",
    "PROGRAM": "Cybersecurity Program Management",
    "RESPONSE": "Incident Response",
    "RISK": "Risk Management",
    "SITUATION": "Situational Awareness",
    "THIRD-PARTIES": "Third-Party Risk Management",
    "THREAT": "Threat and Vulnerability Management",
    "WORKFORCE": "Workforce Management",
}

# Domain keywords for search enrichment
_DOMAIN_KEYWORDS: Dict[str, List[str]] = {
    "ACCESS": [
        "identity",
        "authentication",
        "authorisation",
        "access control",
        "credentials",
        "privilege",
        "MFA",
    ],
    "ARCHITECTURE": [
        "architecture",
        "network design",
        "segmentation",
        "cybersecurity design",
        "zero trust",
    ],
    "ASSET": ["asset inventory", "asset management", "change management", "configuration", "CMDB"],
    "PRIVACY": [
        "privacy",
        "personal information",
        "data protection",
        "APP",
        "NDB",
        "notifiable data breach",
    ],
    "PROGRAM": ["governance", "cybersecurity program", "policy", "strategy", "compliance"],
    "RESPONSE": [
        "incident response",
        "incident management",
        "recovery",
        "forensics",
        "breach response",
    ],
    "RISK": ["risk management", "risk assessment", "risk register", "treatment", "cyber risk"],
    "SITUATION": ["situational awareness", "monitoring", "logging", "threat intelligence", "SIEM"],
    "THIRD-PARTIES": ["third party", "vendor", "supply chain", "outsourcing", "supplier risk"],
    "THREAT": [
        "threat",
        "vulnerability management",
        "patching",
        "penetration testing",
        "threat intelligence",
    ],
    "WORKFORCE": [
        "workforce",
        "training",
        "awareness",
        "personnel",
        "human resources",
        "security culture",
    ],
}

# Practice ID pattern: DOMAIN-MILletter (e.g. ACCESS-1a, PRIVACY-1A)
# or DOMAIN-APnumber (e.g. ACCESS-AP1)
_PRACTICE_RE = re.compile(r"^[A-Z][A-Z0-9-]+-(?:AP\d+|\d+[A-Za-z])$")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _slugify(text: str) -> str:
    """Run slugify."""
    return re.sub(r"[^a-z0-9]+", "_", text.lower()).strip("_")


def _parse_maturity_level(value: object) -> Optional[int]:
    """Extract the integer MIL level from values like ``MIL-2``."""
    if value is None:
        return None
    match = re.search(r"(\d+)", str(value))
    return int(match.group(1)) if match else None


# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------


class AescsfParser(BaseParser):
    """Parse the AESCSF v2 core workbook from the AEMO-hosted Excel workbook.

    Parameters
    ----------
    toolkit_url:
        Direct download URL for the AESCSF Excel file. Defaults to the AEMO
        hosted v2 core workbook.
    include_anti_patterns:
        Whether to include Anti-Pattern practices (Objective IDs ending in ``-AP``).
        Defaults to True.
    """

    def __init__(
        self,
        toolkit_url: str = _CORE_WORKBOOK_URL,
        include_anti_patterns: bool = True,
        **_kwargs,
    ) -> None:
        """Run init."""
        self._toolkit_url = toolkit_url
        self._include_anti_patterns = include_anti_patterns

    # ------------------------------------------------------------------
    # BaseParser implementation
    # ------------------------------------------------------------------

    def parse(self) -> List[RequirementRecord]:
        """Run parse."""
        logger.info("Fetching AESCSF core workbook from %s", self._toolkit_url)
        workbook_bytes = self._fetch_workbook()
        return self._build_records(workbook_bytes)

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _fetch_workbook(self) -> bytes:
        """Run fetch workbook."""
        response = request_with_instrumentation(
            "GET",
            self._toolkit_url,
            logger=logger,
            timeout=60,
            headers={"User-Agent": "aescsf-parser/1.0 (controls ingestion)"},
            system="aemo",
            operation="download_aescsf_workbook",
            request_callable=requests.get,
        )
        response.raise_for_status()
        return response.content

    def _build_records(self, workbook_bytes: bytes) -> List[RequirementRecord]:
        """Run build records."""
        try:
            import openpyxl  # noqa: PLC0415
        except ImportError as exc:
            raise RuntimeError(
                "openpyxl is required for AESCSF parsing. " "Install with: pip install openpyxl"
            ) from exc

        wb = openpyxl.load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)

        if not wb.sheetnames:
            raise RuntimeError("AESCSF workbook has no sheets")

        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise RuntimeError("AESCSF workbook is empty")

        records: List[RequirementRecord] = []

        for row in rows[1:]:
            (
                domain_value,
                objective_id,
                objective,
                practice_id,
                practice,
                guidance,
                mil_value,
                security_profile,
                *_rest,
            ) = row

            domain_slug = str(domain_value).strip() if domain_value is not None else ""
            objective_id_text = str(objective_id).strip() if objective_id is not None else ""
            objective_text = str(objective).strip() if objective is not None else ""
            practice_id_text = str(practice_id).strip() if practice_id is not None else ""
            practice_text = str(practice).strip() if practice is not None else ""
            guidance_text = str(guidance).strip() if guidance is not None else ""
            security_profile_text = (
                str(security_profile).strip() if security_profile is not None else ""
            )

            if not _PRACTICE_RE.match(practice_id_text) or not practice_text:
                continue

            is_anti_pattern = objective_id_text.endswith("-AP") or "-AP" in practice_id_text
            if is_anti_pattern and not self._include_anti_patterns:
                continue

            domain_name = _DOMAIN_NAMES.get(domain_slug, domain_slug.title())
            mil_level = _parse_maturity_level(mil_value)
            control_family = f"{domain_name} – {objective_text}" if objective_text else domain_name
            source_section = f"{domain_name} > {objective_text}" if objective_text else domain_name

            keywords = list(_DOMAIN_KEYWORDS.get(domain_slug, []))
            keywords.extend([domain_slug.lower(), domain_name.lower()])
            if objective_id_text:
                keywords.append(objective_id_text.lower())
            if security_profile_text:
                keywords.append(security_profile_text.lower())
            if mil_level is not None:
                keywords.append(f"mil-{mil_level}")
            if is_anti_pattern:
                keywords.append("anti-pattern")
            keywords = filter_keywords(keywords)

            records.append(
                RequirementRecord(
                    requirement_id=f"AESCSF-{practice_id_text}",
                    framework=FRAMEWORK,
                    framework_version=FRAMEWORK_VERSION,
                    control_family=control_family,
                    maturity_level=mil_level,
                    requirement_text=practice_text,
                    guidance_text=guidance_text,
                    keywords=keywords,
                    source_uri=SOURCE_URI,
                    source_section=source_section,
                    effective_date=EFFECTIVE_DATE,
                    jurisdiction_or_scope=JURISDICTION,
                )
            )

        logger.info(
            "AESCSF: parsed %d practices from core workbook",
            len(records),
        )
        return records
